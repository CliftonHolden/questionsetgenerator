from flask import Flask,render_template,flash,jsonify
from flask import request,redirect,url_for,session,logging,g
import sqlite3
from wtforms import Form,StringField,TextAreaField,PasswordField,validators,SelectField,FileField
from flask_wtf import FlaskForm
from passlib.hash import sha256_crypt
from functools import wraps
import openpyxl,re,os
from flask_wtf.file import FileAllowed, FileRequired
from werkzeug.utils import secure_filename

UPLOAD_FOLDER = os.getcwd()+'\\uploads'
#XL_FILE = "xl_data.xlsx"

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER




#Config sqlite3
DATABASE = "data.db"

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE)
        db.row_factory = sqlite3.Row
    return db

def query_db(query, args=(), one=False):
    cur = get_db().execute(query, args)
    rv = cur.fetchall()
    cur.close()
    return (rv[0] if rv else None) if one else rv

def change_db(query,args=()):
    cur = get_db().execute(query, args)
    get_db().commit()
    cur.close()

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()
#sqlite config extends

#Defining Logged in Wraps
def is_logged_in(f):
    @wraps(f)
    def wrap(*args,**kwargs):
        if'logged_in' in session:
            return f(*args,**kwargs)
        else:
            flash('Unauthorized. Please login','danger')
            return redirect(url_for('login'))
    return wrap
#End of logged in wraps

#Loading data from Excel to sqlite3
def slugify(text, lower=1):
    if lower == 1:
        text = text.strip().lower()
    text = re.sub(r'[^\w _-]+', '', text)
    text = re.sub(r'[- ]+', '_', text)
    return text

def importxl(XL_FILE):
    wb = openpyxl.load_workbook(filename = XL_FILE)
    con = sqlite3.connect('data.db')
    sheets = wb.get_sheet_names()
    #app.logger.info(sheets)
    for sheet in sheets:
        columns=[]
        tup = []
        ws= wb[sheet]
        for rows in ws.iter_rows(min_row=1,max_row=1,min_col=1,max_col=ws.max_column):
            for cell in rows:
                columns.append(cell.value)
        for i,rows in enumerate(ws):
            tuprow=[]
            if i == 0:
                continue
            for row in rows:
                tuprow.append(str(row.value).strip()) if str(row.value).strip() != 'None' else tuprow.append('')
            tup.append(tuple(tuprow))
        insQuery1 = 'INSERT INTO ' + str(slugify(sheet)) + '('
        insQuery2 = ''
        for col in columns:
            insQuery1 += col + ', '
            insQuery2 += '?, '
        insQuery1 = insQuery1[:-2] + ') VALUES('
        insQuery2 = insQuery2[:-2] + ')'
        insQuery = insQuery1 + insQuery2

        con.executemany(insQuery, tup)
        con.commit()

    con.close()
#End of Excel to sqlite3

#Adding Application Routes
@app.route('/')
def index():
    return render_template('home.html')


class RegisterForm(Form):
    name = StringField('Name',[
        validators.DataRequired(),
        validators.Length(min=1, max=50)
    ])
    username = StringField('UserName',[
        validators.Length(min=4,max=25),
        validators.DataRequired()
    ])
    email = StringField('Email',[
        validators.Length(min=12,max=50),
        validators.DataRequired()
    ])
    password = PasswordField('Password', [
        validators.DataRequired(),
        validators.EqualTo('confirm', message = 'Passwords do not match'),
    ])
    confirm = PasswordField('Confirm Password')

@app.route('/register',methods=['GET','POST'])
def register():
    form  = RegisterForm(request.form)
    if request.method == 'POST' and form.validate():
        reg = request.form.to_dict()
        values = [
            reg["name"],
            reg["email"],
            reg["username"],
            sha256_crypt.encrypt(str(reg["password"]))
        ]
        change_db("INSERT INTO users(name,email,username,password) VALUES (?,?,?,?)",values)
        flash("You are registered and can login","success")
        return redirect(url_for('login'))
    return render_template('register.html',form=form)


@app.route('/login',methods=['GET','POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password_candidate = request.form['password']
        result = query_db("SELECT * from users where username=? ",[username],one=True)
        if result is None:
            #flash("Invalid User","danger")
            error = "Invalid Login"
            return render_template('login.html',error=error)
        else:
            app.logger.info('result is > 0')
            password = result['password']
            if sha256_crypt.verify(password_candidate,password):
                app.logger.info('PASSWORD MATCHED')
                session['logged_in'] = True
                session['username'] = username
                session['name'] = result['name']
                session['email'] = result['email']
                return redirect(url_for('dashboard'))
            else:
                error = "Invalid Login"
            return render_template('login.html',error=error)
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    msg = "Successfully Logged out"
    return render_template('login.html',msg = msg)

@app.route('/dashboard')
@is_logged_in
def dashboard():
    return render_template('dashboard.html')
    #Application Route ends

class QuestionSetForm(Form):
    subject = SelectField('Subject',choices=[])
    chapter = SelectField('Chapter',choices=[])

@app.route('/list',methods=['GET','POST'])
@is_logged_in
def list():
    form = QuestionSetForm()
    #sub_choices = [("default","---")]
    sub_query = query_db('SELECT distinct subject,subject from qstnset')
    chap_choices = query_db('SELECT distinct chapter,chapter from qstnset')
    form.subject.choices=[("default","---")]+sub_query
    form.chapter.choices=[("default","---")]+chap_choices
    results = []
    if request.method=='POST':
        search = request.form.to_dict()
        #chap = query_db('SELECT chapter FROM subchaps where rowid = ?',[search["chapter"]])
        values = [
            search["subject"],
            search["chapter"]
        ]
        #app.logger.info(str(chap[0]))
        app.logger.info(search["chapter"])
        #app.logger.info('SEARCHING FOR SPECIFIC SUBJECT AND CHAPTER')
        results = query_db('SELECT * FROM qstnset WHERE subject = ? AND chapter = ?',values)
    return render_template('list.html',form=form,results=results)

@app.route('/get_chapter/<string:subject>')
@is_logged_in
def get_chapter(subject):
    app.logger.info(subject)
    chapters = query_db('SELECT distinct chapter as chapterid,chapter FROM qstnset WHERE subject = ?',[subject])
    chapterArray=[]

    for chapter in chapters:
        chapterObj={}
        chapterObj['chapterid']=chapter['chapterid']
        chapterObj['chapter'] =chapter['chapter']
        chapterArray.append(chapterObj)
    return jsonify({'chapters':chapterArray})

class CreateNewQuestionSetForm(Form):
    question = TextAreaField('Question',[
        validators.DataRequired(),
        validators.Length(min=1, max=10000)
    ])
    subject = StringField('Subject',[
        validators.DataRequired(),
        validators.Length(min=1, max=50)
    ])
    chapter = StringField('Chapter',[
        validators.DataRequired(),
        validators.Length(min=1, max=50)
    ])
    topic = StringField('Topic',[
        validators.DataRequired(),
        validators.Length(min=1, max=50)
    ])
    difficulty = StringField('Difficulty')
    marks = StringField('Marks')
    answer = StringField('Answer',[
        validators.DataRequired(),
        validators.Length(min=1, max=1000)
    ])
    option1 = StringField('Option1',[
        validators.DataRequired(),
        validators.Length(min=1, max=1000)
    ])
    option2 = StringField('Option2',[
        validators.DataRequired(),
        validators.Length(min=1, max=1000)
    ])
    option3 = StringField('Option3',[
        validators.DataRequired(),
        validators.Length(min=1, max=1000)
    ])
    option4 = StringField('Option4',[
        validators.DataRequired(),
        validators.Length(min=1, max=1000)
    ])

@app.route('/create',methods=['GET','POST'])
@is_logged_in
def create():
    form = CreateNewQuestionSetForm(request.form)
    if request.method == 'POST' and form.validate():
        qstnset = request.form.to_dict()
        values = [
            qstnset["question"],
            qstnset["subject"] ,
            qstnset["chapter"],
            qstnset["topic"],
            qstnset["difficulty"],
            qstnset["marks"],
            qstnset["answer"],
            qstnset["option1"],
            qstnset["option2"],
            qstnset["option3"],
            qstnset["option4"]
        ]
        change_db("INSERT INTO qstnset(question,subject,chapter,topic,difficulty,marks,answer,option1,option2,option3,option4) VALUES (?,?,?,?,?,?,?,?,?,?,?)",values)
        #flash("SUCCESSFULLY ENTERED INTO DB","success")
        return redirect(url_for('list'))
    return render_template('create.html',form=form)

@app.route('/edit/<int:id>',methods=['GET','POST'])
@is_logged_in
def edit(id):
    app.logger.info(id)
    form = CreateNewQuestionSetForm(request.form)
    qstn = query_db('SELECT * from qstnset where id = ?',[id])
    app.logger.info(qstn[0])
    for q in qstn:
        form.question.data = q['question']
        form.subject.data = q['subject']
        form.chapter.data = q['chapter']
        form.topic.data = q['topic']
        form.difficulty.data = q['difficulty']
        form.marks.data = q['marks']
        form.answer.data = q['answer']
        form.option1.data = q['option1']
        form.option2.data = q['option2']
        form.option3.data = q['option3']
        form.option4.data = q['option4']

    if request.method == 'POST' and form.validate():
        qstnset = request.form.to_dict()
        values = [
            qstnset["question"],
            qstnset["subject"] ,
            qstnset["chapter"],
            qstnset["topic"],
            qstnset["difficulty"],
            qstnset["marks"],
            qstnset["answer"],
            qstnset["option1"],
            qstnset["option2"],
            qstnset["option3"],
            qstnset["option4"],
            id
        ]
        change_db("UPDATE qstnset SET question = ?, subject = ?, chapter = ?, topic = ?, difficulty = ?, marks = ?, answer = ?, option1 = ?,option2 = ?, option3 = ?, option4 = ? WHERE id = ?",values)
        return redirect(url_for('list'))
    return render_template('edit.html',form=form)

@app.route('/delete/<int:id>',methods=['GET','POST'])
@is_logged_in
def delete(id):
    change_db("DELETE FROM qstnset WHERE id=?",[id])
    return render_template('delete.html')

class BulkUpload(FlaskForm):
    upload = FileField('Upload Question Set',validators=[
        validators.DataRequired(),
        FileAllowed(['xls','xlsx'],'Excel File Only!')
    ])

@app.route('/bulk_upload',methods=['GET','POST'])
@is_logged_in
def bulk_upload():
    form = BulkUpload()
    app.logger.info(form.errors)
    if request.method == 'POST' and form.validate_on_submit():
        #form = BulkUpload(request.form)
        f = form.upload.data

        filename = secure_filename(f.filename)
        f.save(os.path.join(app.config['UPLOAD_FOLDER'],filename))
        XL_FILE = UPLOAD_FOLDER+"\\"+filename
        importxl(XL_FILE)
        return redirect(url_for('list'))
    return render_template('bulkupload.html',form=form)

if __name__ == '__main__':
    app.secret_key='123456'
    app.run(debug=True)
